#
#indexes |  Col 1 |  Col 2 ------------------------------------- >   Col (3*N+1) | Col (3*N+2) ----- > Col (4*N+1)  | Col (4*N+2)  | Col (4*N+3) ------------------------------------- >  Col (7*N+2)  | Col (7*N+3)  ------------------------------------------------------------------------------------------------------------------ >  Col (7*N+6*M+2) |          Col (7*N+6*M+3)          | Col (7*N+6*M+4) -------------------------------------------------------------- >  Col (7*N+N*M+6*M+3)|     
#        ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Row 1  |  Date  | (Robot SN)                                         | ....... |  Site                            |     Date     | (Robot SN)                                               |.......| (Site Name)'s Averages                                                                                                              | ............ | Overall Average acres / hr /robot | Chart's Robots in Sites                                                                              |
# Row 2  |   ---  |   Acres  | First gtg (Time) | Stop request (Time)  | ....... |  (1'st Robot SN) | ............  |      ----    | (Robot SN)'s total acres  | Available time | (Robot SN)  |.......| #Robots | Avg acres/bot | Avg available time | Average acres / hr /robot | (Site Name)'s Total | Overall average  acres / bot / day | ............ |                ---                | (1'st Robot SN)'s total acres in (1'st Site Name) | (N'nd Robot SN)'s total acres in (M'nd Site Name)|
# Row 3  | ....

import csv
import os
from openpyxl import Workbook

def readFile(filename):  
    Robot_col = []
    Site_col = []
    Cumulative_Date = ""

    # Try UTF-8 first, then fallback to Windows encoding if needed
    try:
        f = open(filename, newline='', encoding='utf-8')
        reader = csv.reader(f)
        rows = list(reader)
    except UnicodeDecodeError:
        f = open(filename, newline='', encoding='utf-8-sig', errors='ignore')
        reader = csv.reader(f)
        rows = list(reader)
    finally:
        f.close()

    # Skip first row (header)
    for i, row in enumerate(rows[1:], start=2):
        if len(row) < 3:
            continue

        # Column 1 (non-empty)
        if row[0].strip():
            Robot_col.append(row[0].strip())

        # Column 2 (non-empty)
        if row[1].strip():
            Site_col.append(row[1].strip())

        # Column 3 (second cell only)
        if i == 2:
            Cumulative_Date = row[2].strip() if row[2].strip() else None
    Cumulative_Date = "Cumulative Acres since " + Cumulative_Date

    return Robot_col,Site_col,Cumulative_Date

def get_column_letter(n: int) -> str:
    """
    Convert a positive integer to a Google Sheets column letter (A, B, ..., Z, AA, AB, ...).
    Supports columns 1 through 18278 (A–ZZZ).

    Args:
        n (int): Column number (1-indexed).

    Returns:
        str: Column letter name.

    Raises:
        ValueError: If n is not within 1–18278.
    """
    if not isinstance(n, int) or n < 1 or n > 18278:
        raise ValueError("Column number must be an integer between 1 and 18278.")

    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

# Reading Robots SN & Sites Name & Cumulative Date
filename = "User Inputs.csv"
Robots_SN,Sites_Name,Cumulative_Acres_since_Date  = readFile(filename)
N = len(Robots_SN)
M = len(Sites_Name)

print("Robots_SN List: ",Robots_SN)
print("Robots_SN List length: ",N)

print("Sites_Name List: ",Sites_Name)
print("Sites_Name List length: ",M)

print(Cumulative_Acres_since_Date)

# Check Boundaries 
# ( 7 * N+ N * M + 6 * M =< 18275 ) or M=0 or N=0
if (7 * N+ N * M + 6 * M > 18275) or M == 0 or N == 0 :
    print("Boundaries was broken!")
    print("Can not generate final file according to large number of Robots or Site")
    print("please fit the Boundary equation: 7 * N+ N * M + 6 * M =< 18275 ")
    exit()

# Compute the list length
last_column_index = 7 * N + N * M + 6 * M + 3
print("last_column_index @: ",get_column_letter(last_column_index))

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Filling First row
# Create a list of empty strings of that length
row_1 = [""] * last_column_index
#filling Date
row_1[0] = "Date"
row_1[4 * N + 1] = "Date"
row_1[3 * N + 1] = "Site"

#filling Robots SN 
if N != 0:
    robotSummaryStartIndex = 4 * N + 2
    for index, Robot_SN in enumerate(Robots_SN):
        row_1[1 + index * 3] = "SN " + Robot_SN
        row_1[robotSummaryStartIndex + index * 3] = "SN " + Robot_SN

#filling Sites Name 
if M != 0:
    siteSummaryStartIndex = 7 * N + 2
    for index, Site_Name in enumerate(Sites_Name):
        row_1[siteSummaryStartIndex + index * 6] = Site_Name + "'s Averages"

# Overall Average acres / hr /robot
row_1[7 * N + 6 * M + 2] = "Overall Average acres / hr /robot"

# filling Chart's Robots in Sites 
row_1[7 * N + 6 * M + 3] = "Chart's Robots in Sites"



#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Filling Second row
# Create a list of empty strings of that length
row_2 = [""] * last_column_index

#filling Robots SN 
if N != 0:
    robotSummaryStartIndex = 4 * N + 2
    robotSitesStartIndex = 3 * N + 1
    for index, Robot_SN in enumerate(Robots_SN):
        row_2[1 + index * 3] = "Acres"
        row_2[2 + index * 3] = "First gtg (Time)"
        row_2[3 + index * 3] = "Stop request (Time)"
        row_2[robotSitesStartIndex + index] = "SN " + Robot_SN
        row_2[robotSummaryStartIndex + index * 3] = "SN " + Robot_SN + "'s total acres"
        row_2[robotSummaryStartIndex + 1 + index * 3] = "Available time"
        row_2[robotSummaryStartIndex + 2 + index * 3] = "SN " + Robot_SN 

#filling Sites Name 
if M != 0:
    siteSummaryStartIndex = 7 * N + 2
    for index, Site_Name in enumerate(Sites_Name):
        row_2[siteSummaryStartIndex + index * 6] = "#Robots"
        row_2[siteSummaryStartIndex + 1 + index * 6] = "Avg acres/bot"
        row_2[siteSummaryStartIndex + 2 + index * 6] = "Avg available time"
        row_2[siteSummaryStartIndex + 3 + index * 6] = "Average acres / hr /robot"
        row_2[siteSummaryStartIndex + 4 + index * 6] = Site_Name + "'s Total"
        row_2[siteSummaryStartIndex + 5 + index * 6] = "Overall average acres / bot / day"


# filling Chart's Robots in Sites 
if M != 0 and N != 0:
    ChartStartIndex = 7 * N + 6 * M + 3
    for Site_Name in Sites_Name:
        for Robot_SN in Robots_SN:
            row_2[ChartStartIndex] = Robot_SN + "'s total acres in " + Site_Name
            ChartStartIndex = ChartStartIndex + 1



#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Filling third row
# References Indexes
Robot_Data_Acres_start_index =  2   #loop: + 3 * index (0,1,....,N-1)
Robot_Data_First_gtg_Time_start_index =  3   #loop: + 3 * index (0,1,....,N-1)
Robot_Data_Stop_request_Time_start_index =  4   #loop: + 3 * index (0,1,....,N-1)

Robot_Data_Site_start_index =  3*N+2   #loop: + 1 * index (0,1,....,N-1)

Date_index = 4*N+2

Robot_total_acres_start_index =  4*N+3   #loop: + 3 * index (0,1,....,N-1)
Robot_Available_time_start_index =  4*N+4   #loop: + 3 * index (0,1,....,N-1)
Robot_SN_start_index =  4*N+5   #loop: + 3 * index (0,1,....,N-1)

Site_Number_Robots_start_index = 7*N+3      #loop: + 6 * index (0,1,....,M-1)
Site_Avg_acres_bot_start_index = 7*N+4      #loop: + 6 * index (0,1,....,M-1)
Site_Avg_available_time_start_index = 7*N+5     #loop: + 6 * index (0,1,....,M-1)
Site_Average_acres_hr_robot_start_index = 7*N+6     #loop: + 6 * index (0,1,....,M-1)
Site_Total_acres_start_index = 7*N+7        #loop: + 6 * index (0,1,....,M-1)
Site_Overall_average_acres_bot_day_start_index = 7*N+8      #loop: + 6 * index (0,1,....,M-1)

Overall_Average_acres_hr_robot =7*N+6*M+3

# Create a list of empty strings of that length
row_3 = [""] * last_column_index

#filling Date
repeated_date_column = get_column_letter(Date_index)
row_3[4 * N + 1] ='=IF('+repeated_date_column+'2="Overall average","' + Cumulative_Acres_since_Date + '",IF('+repeated_date_column+'2="' + Cumulative_Acres_since_Date + '",,IF(AND(ISBLANK(A3),ISBLANK(A2)),,IF(ISBLANK(A3),"Overall average",A3))))'
#filling Robots SN 
if N != 0:
    robotSummaryStartIndex = 4 * N + 2
    for index, Robot_SN in enumerate(Robots_SN):
        repeated_column_1 = get_column_letter(Robot_total_acres_start_index + 3 * index)
        row_3[robotSummaryStartIndex + index * 3] = '=IF(OR('+repeated_date_column+'3="Overall average",ISBLANK('+repeated_date_column+'3)),,IF('+repeated_date_column+'3="' + Cumulative_Acres_since_Date + '",SUM($'+ repeated_column_1 +'$3 :INDEX(' + repeated_column_1 + ':' + repeated_column_1 + ',ROW()-1)),'+ get_column_letter(Robot_Data_Acres_start_index+ 3 * index) +'3))'

        start_repeated_column = get_column_letter(Robot_Data_First_gtg_Time_start_index+ 3 * index)
        end_repeated_column = get_column_letter(Robot_Data_Stop_request_Time_start_index+ 3 * index)
        row_3[robotSummaryStartIndex + 1 + index * 3] = '=IF(OR('+repeated_date_column+'3="' + Cumulative_Acres_since_Date + '",'+repeated_date_column+'3="Overall average",ISBLANK('+repeated_date_column+'3),ISBLANK(' + start_repeated_column + '3), ISBLANK(' + end_repeated_column + '3)),,IF(' + end_repeated_column + '3 < ' + start_repeated_column + '3, HOUR(ABS(' + end_repeated_column + '3-' + start_repeated_column + '3+1)) + (MINUTE(ABS(' + end_repeated_column + '3-' + start_repeated_column + '3+1))/60), HOUR(ABS(' + end_repeated_column + '3-' + start_repeated_column + '3)) + (MINUTE(ABS(' + end_repeated_column + '3-' + start_repeated_column + '3))/60)))'

        repeated_column = get_column_letter(Robot_Available_time_start_index+ 3 * index)        
        row_3[robotSummaryStartIndex + 2 + index * 3] = '=IF(ISBLANK('+repeated_column+'3),,'+get_column_letter(Robot_total_acres_start_index + 3 * index)+'3/'+repeated_column+'3)'

#filling Sites Name 
if M != 0:
    siteSummaryStartIndex = 7 * N + 2
    for index, Site_Name in enumerate(Sites_Name):

        all_Robots_Site = 'IF(' + get_column_letter(Robot_Data_Site_start_index) +'3="'+ Site_Name +'",1,0)'
        for i in range(1,N):
            all_Robots_Site += '+IF('+ get_column_letter(Robot_Data_Site_start_index  + 1 * i) +'3="'+ Site_Name +'",1,0)'
        row_3[siteSummaryStartIndex + index * 6] = '=IF(OR(' + repeated_date_column + '3="' + Cumulative_Acres_since_Date + '",' + repeated_date_column + '3="Overall average",ISBLANK(' + repeated_date_column + '3)),,' + all_Robots_Site +')'

        all_Robots_Acres = 'IF(' + get_column_letter(Robot_Data_Site_start_index) +'3="'+ Site_Name +'",' + get_column_letter(Robot_total_acres_start_index) +'3,0)'
        for i in range(1,N):
            all_Robots_Acres += '+IF('+ get_column_letter(Robot_Data_Site_start_index  + 1 * i) +'3="'+ Site_Name +'",' + get_column_letter(Robot_total_acres_start_index + 3 * i) +'3,0)'
        repeated_column_1 = get_column_letter(Site_Number_Robots_start_index+ 6 * index)
        row_3[siteSummaryStartIndex + 1 + index * 6] = '=IF(OR(' + repeated_date_column + '3="' + Cumulative_Acres_since_Date + '",' + repeated_date_column + '3="Overall average",ISBLANK(' + repeated_date_column + '3)),,IF('+repeated_column_1+'3=0,0,(' + all_Robots_Acres + ')/'+repeated_column_1+'3))'

        all_Robots_Ava_time = 'IF(' + get_column_letter(Robot_Data_Site_start_index) +'3="'+ Site_Name +'",' + get_column_letter(Robot_Available_time_start_index) +'3,0)'
        for i in range(1,N):
            all_Robots_Ava_time += '+IF('+ get_column_letter(Robot_Data_Site_start_index  + 1 * i) +'3="'+ Site_Name +'",' + get_column_letter(Robot_Available_time_start_index + 3 * i) +'3,0)'
        repeated_column_1 = get_column_letter(Site_Number_Robots_start_index+ 6 * index)
        row_3[siteSummaryStartIndex + 2 + index * 6] = '=IF(OR(' + repeated_date_column + '3="' + Cumulative_Acres_since_Date + '",' + repeated_date_column + '3="Overall average",ISBLANK(' + repeated_date_column + '3)),,IF('+repeated_column_1+'3=0,0,(' + all_Robots_Ava_time + ')/'+repeated_column_1+'3))'

        repeated_column_1 = get_column_letter(Site_Average_acres_hr_robot_start_index + 6 * index)
        repeated_column_2 = get_column_letter(Site_Avg_available_time_start_index + 6 * index)
        row_3[siteSummaryStartIndex + 3 + index * 6] = '=IF(OR(' + repeated_date_column + '3="' + Cumulative_Acres_since_Date + '",ISBLANK(' + repeated_date_column + '3)),,IF('+repeated_date_column+'3="Overall average",AVERAGE($'+repeated_column_1+'$3 :INDEX('+repeated_column_1+':'+repeated_column_1+',ROW()-1)),IF('+repeated_column_2+'3=0,0,'+get_column_letter(Site_Avg_acres_bot_start_index + 6 * index)+'3/'+repeated_column_2+'3)))'

        all_Robots_Acres = 'IF(' + get_column_letter(Robot_Data_Site_start_index) +'3="'+ Site_Name +'",' + get_column_letter(Robot_total_acres_start_index) +'3,0)'
        for i in range(1,N):
            all_Robots_Acres += '+IF('+ get_column_letter(Robot_Data_Site_start_index  + 1 * i) +'3="'+ Site_Name +'",' + get_column_letter(Robot_total_acres_start_index + 3 * i) +'3,0)'
        repeated_column_1 = get_column_letter(Site_Total_acres_start_index + 6 * index)
        row_3[siteSummaryStartIndex + 4 + index * 6] = '=IF(OR(ISBLANK(' + repeated_date_column + '3), TRIM(' + repeated_date_column + '3)="Overall average"),,IF(' + repeated_date_column + '3="' + Cumulative_Acres_since_Date + '",SUM($' + repeated_column_1 + '$3 :INDEX(' + repeated_column_1 + ':' + repeated_column_1 + ',ROW()-1)),(' + all_Robots_Acres + ')))'

        repeated_column_1 = get_column_letter(Site_Avg_acres_bot_start_index + 6 * index)
        row_3[siteSummaryStartIndex + 5 + index * 6] = '=IF('+ repeated_date_column +'3="Overall average",AVERAGE($'+ repeated_column_1 +'$3 :INDEX('+ repeated_column_1 +':'+ repeated_column_1 +',ROW()-1)),)'

# Overall Average acres / hr /robot
repeated_column_1 = get_column_letter(Overall_Average_acres_hr_robot)
all_Sites = get_column_letter(Site_Average_acres_hr_robot_start_index) +'3'
for index in range(1,M):
    all_Sites += ','+ get_column_letter(Site_Average_acres_hr_robot_start_index  + 6 * index) +'3'
row_3[7 * N + 6 * M + 2] = '=IF(OR(' + repeated_date_column + '3="' + Cumulative_Acres_since_Date + '",ISBLANK(' + repeated_date_column + '3)),,IF(' + repeated_date_column + '3="Overall average",AVERAGE($' + repeated_column_1 + '$3 :INDEX(' + repeated_column_1 + ':' + repeated_column_1 + ',ROW()-1)),AVERAGE('+all_Sites+')))'

# filling Chart's Robots in Sites 
if M != 0 and N != 0:
    ChartStartIndex = 7 * N + 6 * M + 3
    for Site_Name in Sites_Name:
        for index, Robot_SN in enumerate(Robots_SN):
            repeated_column_1 = get_column_letter(ChartStartIndex +1)
            row_3[ChartStartIndex] = '=IF(OR(' + repeated_date_column + '3="Overall average",ISBLANK(' + repeated_date_column + '3)),,IF('+ repeated_date_column +'3="' + Cumulative_Acres_since_Date + '",SUM($'+ repeated_column_1 +'$3 :INDEX('+ repeated_column_1 +':'+ repeated_column_1 +',ROW()-1)),IF('+ get_column_letter(Robot_Data_Site_start_index + 1 * index) +'3 = "'+ Site_Name +'",'+ get_column_letter(Robot_total_acres_start_index + 3 * index) +'3,0)))'
            ChartStartIndex = ChartStartIndex + 1

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------

# print(row_1)
# print(row_2)
# for idx,cell in enumerate(row_3):
#     print("The Column #( ",get_column_letter(idx +1)," ): ",cell)

wb = Workbook()
ws = wb.active

# Write both rows
ws.append(row_1)
ws.append(row_2)
ws.append(row_3)

# Save to Excel file
output = str(N) +"_Robots_"+ str(M) +"_Sites.xlsx"

# Check if file exists, then delete it
if os.path.exists(output):
    os.remove(output)
    print("Old file deleted:", output)

wb.save(output)

print("Excel file created successfully: ",output)
 
